"""
Excel Exporter Module for converting structured JSON to Excel format
"""

import json
import logging
import os
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional

try:
    import pandas as pd
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
except ImportError as e:
    pd = None
    logging.warning(f"Required packages not found: {e}. Please install pandas and openpyxl.")


class MenuExcelExporter:
    """
    Exporter class for converting structured menu JSON to Excel format
    """
    
    def __init__(self):
        """Initialize the Excel exporter"""
        if pd is None:
            raise ImportError("pandas and openpyxl packages are required for Excel export")
    
    def json_to_excel(
        self, 
        json_data: Dict, 
        output_path: str,
        include_metadata: bool = True,
        single_sheet: bool = True
    ) -> bool:
        """
        Convert structured JSON data to Excel file with single or multiple sheets
        
        Args:
            json_data: The structured menu JSON data
            output_path: Path to save the Excel file
            include_metadata: Whether to include metadata sheet
            single_sheet: Whether to create single sheet (True) or multiple sheets (False)
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            # Create Excel writer
            with pd.ExcelWriter(output_path, engine='openpyxl') as writer:
                
                if single_sheet:
                    # Export all data to a single sheet
                    self._export_single_sheet(json_data, writer)
                else:
                    # Export Restaurant sheet
                    self._export_restaurant_sheet(json_data, writer)
                    
                    # Export Categories sheet
                    self._export_categories_sheet(json_data, writer)
                    
                    # Export Items sheet
                    self._export_items_sheet(json_data, writer)
                    
                    # Export AddOnGroups sheet
                    self._export_addongroups_sheet(json_data, writer)
                    
                    # Export metadata sheet if requested
                    if include_metadata:
                        self._export_metadata_sheet(json_data, writer)
            
            # Apply formatting
            self._apply_formatting(output_path)
            
            logging.info(f"Successfully exported menu data to Excel: {output_path}")
            return True
            
        except Exception as e:
            logging.error(f"Failed to export to Excel: {e}")
            return False
    
    def _export_restaurant_sheet(self, json_data: Dict, writer) -> None:
        """Export restaurant information to Excel sheet"""
        restaurant_data = json_data.get("restaurant", {})
        
        df = pd.DataFrame([{
            "Name": restaurant_data.get("name", "Unknown"),
            "Source Image": restaurant_data.get("source_image", "Unknown"),
            "Export Date": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
            "Total Categories": len(json_data.get("categories", [])),
            "Total Items": len(json_data.get("items", [])),
            "Total Addon Groups": len(json_data.get("addongroups", []))
        }])
        
        df.to_excel(writer, sheet_name="Restaurant", index=False)
    
    def _export_single_sheet(self, json_data: Dict, writer) -> None:
        """Export all menu data to a single Excel sheet matching data_reference.json schema"""
        restaurant_data = json_data.get("restaurant", {})
        restaurant_name = restaurant_data.get("restaurantname", "Unknown")
        
        # Get areas (use first area if available)
        areas = json_data.get("areas", [])
        area_id = areas[0].get("areaid") if areas else None
        area_display_name = areas[0].get("displayname") if areas else None
        
        # Create category lookup
        category_map = {cat.get("categoryid"): cat for cat in json_data.get("categories", [])}
        
        # Create addon group lookup
        addon_group_map = {grp.get("addongroupid"): grp for grp in json_data.get("addongroups", [])}
        
        rows = []
        items = json_data.get("items", [])
        
        if not items:
            # Create empty row with just restaurant info
            rows.append(self._create_empty_row_new_schema(restaurant_name, area_id, area_display_name))
        else:
            # Process each item
            for item in items:
                rows.extend(self._create_item_rows(
                    item,
                    restaurant_name,
                    area_id,
                    area_display_name,
                    category_map,
                    addon_group_map
                ))
        
        # Create DataFrame with exact column order
        df = pd.DataFrame(rows, columns=[
            "restaurant_name",
            "area_id",
            "area_display_name",
            "category_id",
            "category_name",
            "category_image_url",
            "category_timings",
            "category_rank",
            "item_id",
            "item_name",
            "item_description",
            "price",
            "rank",
            "image_url",
            "instock",
            "variation_item_id",
            "variation_id",
            "variation_name",
            "variation_price",
            "addon_name",
            "addon_item_selection",
            "addon_item_selection_min",
            "addon_item_selection_max",
            "addon_price",
            "addon_id",
            "addon_group_id",
            "addon_group_name",
        ])
        
        df.to_excel(writer, sheet_name="Menu_Data", index=False)
    
    def _create_empty_row_new_schema(self, restaurant_name: str, area_id, area_display_name) -> Dict:
        """Create an empty row with restaurant info for new schema"""
        return {
            "restaurant_name": restaurant_name,
            "area_id": area_id,
            "area_display_name": area_display_name,
            "category_id": None,
            "category_name": None,
            "category_image_url": None,
            "category_timings": None,
            "category_rank": None,
            "item_id": None,
            "item_name": None,
            "item_description": None,
            "price": None,
            "rank": None,
            "image_url": None,
            "instock": None,
            "variation_item_id": None,
            "variation_id": None,
            "variation_name": None,
            "variation_price": None,
            "addon_name": None,
            "addon_item_selection": None,
            "addon_item_selection_min": None,
            "addon_item_selection_max": None,
            "addon_price": None,
            "addon_id": None,
            "addon_group_id": None,
            "addon_group_name": None
        }
    
    def _create_item_rows(self, item: Dict, restaurant_name: str, area_id, area_display_name,
                          category_map: Dict, addon_group_map: Dict) -> list:
        """Create all necessary rows for a single item (including variations and addons)"""
        rows = []
        
        # Get category info
        category_id = item.get("item_categoryid")
        category_info = category_map.get(category_id, {})
        
        # Base row data
        base_data = {
            "restaurant_name": restaurant_name,
            "area_id": area_id,
            "area_display_name": area_display_name,
            "category_id": category_id,
            "category_name": category_info.get("categoryname", ""),
            "category_image_url": category_info.get("category_image_url"),
            "category_timings": category_info.get("categorytimings"),
            "category_rank": category_info.get("categoryrank"),
            "item_id": item.get("itemid"),
            "item_name": item.get("itemname", ""),
            "item_description": item.get("itemdescription", ""),
            "price": item.get("price"),
            "rank": item.get("itemrank"),
            "image_url": item.get("item_image_url"),
            "instock": item.get("instock", "2"),
        }
        
        variations = item.get("variation", [])
        addon_refs = item.get("addon", [])
        
        # Determine how to create rows
        if not variations and not addon_refs:
            # Simple item - no variations or addons
            row = base_data.copy()
            row.update({
                "variation_item_id": None,
                "variation_id": None,
                "variation_name": None,
                "variation_price": None,
                "addon_name": None,
                "addon_item_selection": None,
                "addon_item_selection_min": None,
                "addon_item_selection_max": None,
                "addon_price": None,
                "addon_id": None,
                "addon_group_id": None,
                "addon_group_name": None,
            })
            rows.append(row)
        
        elif variations and not addon_refs:
            # Item with variations, no addons
            for variation in variations:
                row = base_data.copy()
                row.update({
                    "variation_item_id": variation.get("variationitemid"),
                    "variation_id": variation.get("variationid"),
                    "variation_name": variation.get("variation_name"),
                    "variation_price": variation.get("variation_price"),
                    "addon_name": None,
                    "addon_item_selection": None,
                    "addon_item_selection_min": None,
                    "addon_item_selection_max": None,
                    "addon_price": None,
                    "addon_id": None,
                    "addon_group_id": None,
                    "addon_group_name": None,
                })
                rows.append(row)
        
        elif not variations and addon_refs:
            # Item with addons, no variations
            for addon_ref in addon_refs:
                addon_group_id = addon_ref.get("addon_group_id")
                addon_group = addon_group_map.get(addon_group_id, {})
                addon_items = addon_group.get("addongroupitems", [])
                
                if addon_items:
                    for addon_item in addon_items:
                        row = base_data.copy()
                        row.update({
                            "variation_item_id": None,
                            "variation_id": None,
                            "variation_name": None,
                            "variation_price": None,
                            "addon_name": addon_item.get("addonitem_name"),
                            "addon_item_selection": addon_ref.get("addon_item_selection"),
                            "addon_item_selection_min": addon_ref.get("addon_item_selection_min", "0"),
                            "addon_item_selection_max": addon_ref.get("addon_item_selection_max", "2"),
                            "addon_price": addon_item.get("addonitem_price"),
                            "addon_id": addon_item.get("addonitemid"),
                            "addon_group_id": addon_group_id,
                            "addon_group_name": addon_group.get("addongroup_name", ""),
                        })
                        rows.append(row)
                else:
                    # Addon group with no items
                    row = base_data.copy()
                    row.update({
                        "variation_item_id": None,
                        "variation_id": None,
                        "variation_name": None,
                        "variation_price": None,
                        "addon_name": None,
                        "addon_item_selection": addon_ref.get("addon_item_selection"),
                        "addon_item_selection_min": addon_ref.get("addon_item_selection_min", "0"),
                        "addon_item_selection_max": addon_ref.get("addon_item_selection_max", "2"),
                        "addon_price": None,
                        "addon_id": None,
                        "addon_group_id": addon_group_id,
                        "addon_group_name": addon_group.get("addongroup_name", ""),
                    })
                    rows.append(row)
        
        else:
            # Item with both variations and addons - create cartesian product
            for variation in variations:
                for addon_ref in addon_refs:
                    addon_group_id = addon_ref.get("addon_group_id")
                    addon_group = addon_group_map.get(addon_group_id, {})
                    addon_items = addon_group.get("addongroupitems", [])
                    
                    if addon_items:
                        for addon_item in addon_items:
                            row = base_data.copy()
                            row.update({
                                "variation_item_id": variation.get("variationitemid"),
                                "variation_id": variation.get("variationid"),
                                "variation_name": variation.get("variation_name"),
                                "variation_price": variation.get("variation_price"),
                                "addon_name": addon_item.get("addonitem_name"),
                                "addon_item_selection": addon_ref.get("addon_item_selection"),
                                "addon_item_selection_min": addon_ref.get("addon_item_selection_min", "0"),
                                "addon_item_selection_max": addon_ref.get("addon_item_selection_max", "2"),
                                "addon_price": addon_item.get("addonitem_price"),
                                "addon_id": addon_item.get("addonitemid"),
                                "addon_group_id": addon_group_id,
                                "addon_group_name": addon_group.get("addongroup_name", ""),
                            })
                            rows.append(row)
                    else:
                        # Variation with addon group that has no items
                        row = base_data.copy()
                        row.update({
                            "variation_item_id": variation.get("variationitemid"),
                            "variation_id": variation.get("variationid"),
                            "variation_name": variation.get("variation_name"),
                            "variation_price": variation.get("variation_price"),
                            "addon_name": None,
                            "addon_item_selection": addon_ref.get("addon_item_selection"),
                            "addon_item_selection_min": addon_ref.get("addon_item_selection_min", "0"),
                            "addon_item_selection_max": addon_ref.get("addon_item_selection_max", "2"),
                            "addon_price": None,
                            "addon_id": None,
                            "addon_group_id": addon_group_id,
                            "addon_group_name": addon_group.get("addongroup_name", ""),
                        })
                        rows.append(row)
        
        return rows
    
    def _create_empty_row(self, restaurant_data: Dict) -> Dict:
        """Create an empty row with just restaurant info"""
        return {
            "restaurant_name": restaurant_data.get("name", "Unknown"),
            "area_id": None,
            "area_display_name": None,
            "category_id": None,
            "category_name": None,
            "category_image_url": None,
            "category_timings": None,
            "category_rank": None,
            "item_id": None,
            "item_name": None,
            "item_description": None,
            "price": None,
            "rank": None,
            "image_url": None,
            "instock": None,
            "variation_item_id": None,
            "variation_id": None,
            "variation_name": None,
            "variation_price": None,
            "addon_name": None,
            "addon_item_selection": None,
            "addon_item_selection_min": None,
            "addon_item_selection_max": None,
            "addon_price": None,
            "addon_id": None,
            "addon_group_id": None,
            "addon_group_name": None
        }
    
    def _find_category_info(self, item: Dict, categories: list) -> Dict:
        """Find category information for an item"""
        category_id = item.get("categoryid")
        category_info = {
            "category_id": category_id,
            "category_name": None,
            "category_rank": None
        }
        
        if category_id:
            for cat in categories:
                if cat.get("categoryid") == category_id:
                    category_info.update({
                        "category_name": cat.get("categoryname", ""),
                        "category_rank": cat.get("rank")
                    })
                    break
        else:
            # If no category ID, try to match by name or assign first category
            if categories:
                first_cat = categories[0]
                category_info.update({
                    "category_id": first_cat.get("categoryid"),
                    "category_name": first_cat.get("categoryname", ""),
                    "category_rank": first_cat.get("rank")
                })
        
        return category_info
    
    def _get_item_addons(self, item: Dict, addongroups: list) -> list:
        """Get addon information for an item"""
        addons = []
        item_addon_groups = item.get("addongroups", [])
        
        for addon_group_id in item_addon_groups:
            for group in addongroups:
                if group.get("group_id") == addon_group_id:
                    for addon_item in group.get("items", []):
                        addons.append({
                            "addon_name": addon_item,
                            "addon_item_selection": 1,
                            "addon_item_selection_min": group.get("min_select", 0),
                            "addon_item_selection_max": group.get("max_select", 2),
                            "addon_price": None,  # Price not available in current schema
                            "addon_id": None,
                            "addon_group_id": group.get("group_id"),
                            "addon_group_name": group.get("group_name", "")
                        })
        
        return addons
    
    def _create_base_row(self, restaurant_data: Dict, category_info: Dict, item: Dict) -> Dict:
        """Create base row data for an item"""
        return {
            "restaurant_name": restaurant_data.get("name", "Unknown"),
            "area_id": None,  # Not available in current schema
            "area_display_name": None,  # Not available in current schema
            "category_id": category_info.get("category_id"),
            "category_name": category_info.get("category_name"),
            "category_image_url": None,  # Not available in current schema
            "category_timings": None,  # Not available in current schema
            "category_rank": category_info.get("category_rank"),
            "item_id": item.get("itemid"),
            "item_name": item.get("itemname", ""),
            "item_description": item.get("description", "").replace('\n', ' ') if item.get("description") else "",
            "price": item.get("price"),
            "rank": None,  # Could be derived from confidence or order
            "image_url": None,  # Not available in current schema
            "instock": item.get("instock", 2),
            "variation_item_id": None,
            "variation_id": None,
            "variation_name": None,
            "variation_price": None,
            "addon_name": None,
            "addon_item_selection": None,
            "addon_item_selection_min": None,
            "addon_item_selection_max": None,
            "addon_price": None,
            "addon_id": None,
            "addon_group_id": None,
            "addon_group_name": None
        }
    
    def _export_categories_sheet(self, json_data: Dict, writer) -> None:
        """Export categories information to Excel sheet"""
        categories = json_data.get("categories", [])
        
        if not categories:
            # Create empty dataframe with headers
            df = pd.DataFrame(columns=[
                "Category Name", "Category ID", "Confidence", 
                "Coordinates", "Rank", "Active"
            ])
        else:
            df = pd.DataFrame([{
                "Category Name": cat.get("categoryname", ""),
                "Category ID": cat.get("categoryid"),
                "Confidence": cat.get("confidence", 1.0),
                "Coordinates": str(cat.get("coordinates")) if cat.get("coordinates") else "",
                "Rank": cat.get("rank"),
                "Active": cat.get("active", "1")
            } for cat in categories])
        
        df.to_excel(writer, sheet_name="Categories", index=False)
    
    def _export_items_sheet(self, json_data: Dict, writer) -> None:
        """Export menu items information to Excel sheet"""
        items = json_data.get("items", [])
        
        if not items:
            # Create empty dataframe with headers
            df = pd.DataFrame(columns=[
                "Item Name", "Item ID", "Category ID", "Description", 
                "Price", "Price Variants", "Currency", "In Stock", 
                "Availability", "Tags", "Addon Groups", "Coordinates", "Confidence"
            ])
        else:
            df = pd.DataFrame([{
                "Item Name": item.get("itemname", ""),
                "Item ID": item.get("itemid"),
                "Category ID": item.get("categoryid"),
                "Description": item.get("description", ""),
                "Price": item.get("price"),
                "Price Variants": ", ".join(map(str, item.get("price_variants", []))),
                "Currency": item.get("currency", "INR"),
                "In Stock": item.get("instock", 2),
                "Availability": item.get("availability", 1),
                "Tags": ", ".join(item.get("tags", [])),
                "Addon Groups": ", ".join(map(str, item.get("addongroups", []))),
                "Coordinates": str(item.get("coordinates")) if item.get("coordinates") else "",
                "Confidence": item.get("confidence", 1.0)
            } for item in items])
        
        df.to_excel(writer, sheet_name="Items", index=False)
    
    def _export_addongroups_sheet(self, json_data: Dict, writer) -> None:
        """Export addon groups information to Excel sheet"""
        addongroups = json_data.get("addongroups", [])
        
        if not addongroups:
            # Create empty dataframe with headers
            df = pd.DataFrame(columns=[
                "Group Name", "Group ID", "Min Select", "Max Select", "Items"
            ])
        else:
            df = pd.DataFrame([{
                "Group Name": group.get("group_name", ""),
                "Group ID": group.get("group_id"),
                "Min Select": group.get("min_select", 0),
                "Max Select": group.get("max_select", 2),
                "Items": ", ".join(group.get("items", []))
            } for group in addongroups])
        
        df.to_excel(writer, sheet_name="AddOnGroups", index=False)
    
    def _export_metadata_sheet(self, json_data: Dict, writer) -> None:
        """Export metadata and audit log to Excel sheet"""
        # Create metadata summary
        metadata = [
            {"Property": "Export Timestamp", "Value": datetime.now().isoformat()},
            {"Property": "JSON Structure Version", "Value": "1.0"},
            {"Property": "Source", "Value": "OCR + Gemini LLM Processing"},
            {"Property": "Categories Count", "Value": len(json_data.get("categories", []))},
            {"Property": "Items Count", "Value": len(json_data.get("items", []))},
            {"Property": "Addon Groups Count", "Value": len(json_data.get("addongroups", []))},
        ]
        
        # Add audit log entries if available
        audit_log = json_data.get("audit_log", [])
        if audit_log:
            for i, entry in enumerate(audit_log):
                metadata.append({
                    "Property": f"Audit Log {i+1}",
                    "Value": str(entry)
                })
        
        df = pd.DataFrame(metadata)
        df.to_excel(writer, sheet_name="Metadata", index=False)
    
    def _apply_formatting(self, excel_path: str) -> None:
        """Apply formatting to the Excel file"""
        try:
            from openpyxl import load_workbook
            
            wb = load_workbook(excel_path)
            
            # Define styles
            header_font = Font(bold=True, color="FFFFFF")
            header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
            center_alignment = Alignment(horizontal="center", vertical="center")
            
            # Format each sheet
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                
                # Format header row
                if ws.max_row > 0:
                    for cell in ws[1]:
                        cell.font = header_font
                        cell.fill = header_fill
                        cell.alignment = center_alignment
                
                # Auto-adjust column widths
                for column in ws.columns:
                    max_length = 0
                    column_letter = column[0].column_letter
                    
                    for cell in column:
                        try:
                            if len(str(cell.value)) > max_length:
                                max_length = len(str(cell.value))
                        except:
                            pass
                    
                    adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                    ws.column_dimensions[column_letter].width = adjusted_width
            
            wb.save(excel_path)
            
        except Exception as e:
            logging.warning(f"Could not apply formatting to Excel file: {e}")
    
    def save_json(self, json_data: Dict, output_path: str) -> bool:
        """
        Save JSON data to file
        
        Args:
            json_data: The structured menu JSON data
            output_path: Path to save the JSON file
        
        Returns:
            bool: True if successful, False otherwise
        """
        try:
            # Ensure output directory exists
            os.makedirs(os.path.dirname(output_path), exist_ok=True)
            
            with open(output_path, 'w', encoding='utf-8') as f:
                json.dump(json_data, f, indent=2, ensure_ascii=False)
            
            logging.info(f"Successfully saved JSON data: {output_path}")
            return True
            
        except Exception as e:
            logging.error(f"Failed to save JSON data: {e}")
            return False


def export_menu_to_excel(
    json_data: Dict,
    output_dir: str,
    base_filename: str = "menu_export",
    include_json: bool = True,
    include_metadata: bool = True,
    single_sheet: bool = True
) -> tuple[str, Optional[str]]:
    """
    Convenience function to export menu JSON to Excel and optionally JSON
    
    Args:
        json_data: The structured menu JSON data
        output_dir: Directory to save the files
        base_filename: Base filename (without extension)
        include_json: Whether to also save as JSON file
        include_metadata: Whether to include metadata sheet in Excel (ignored if single_sheet=True)
        single_sheet: Whether to create single sheet (True) or multiple sheets (False)
    
    Returns:
        tuple: (excel_path, json_path) - json_path is None if include_json is False
    """
    exporter = MenuExcelExporter()
    
    # Generate file paths
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    excel_filename = f"{base_filename}_{timestamp}.xlsx"
    excel_path = os.path.join(output_dir, excel_filename)
    
    json_path = None
    if include_json:
        json_filename = f"{base_filename}_{timestamp}.json"
        json_path = os.path.join(output_dir, json_filename)
    
    # Export to Excel
    excel_success = exporter.json_to_excel(json_data, excel_path, include_metadata, single_sheet)
    if not excel_success:
        raise Exception("Failed to export to Excel")
    
    # Save JSON if requested
    if include_json:
        json_success = exporter.save_json(json_data, json_path)
        if not json_success:
            logging.warning("Failed to save JSON file, but Excel export succeeded")
            json_path = None
    
    return excel_path, json_path


# Example usage and testing
if __name__ == "__main__":
    # Test with sample data
    sample_json = {
        "restaurant": {
            "name": "Test Calzone Restaurant",
            "source_image": "sample/menu.jpg"
        },
        "categories": [
            {
                "categoryname": "Calzone Menu",
                "categoryid": None,
                "confidence": 1.0,
                "coordinates": None,
                "rank": None,
                "active": "1"
            }
        ],
        "items": [
            {
                "itemname": "Chicken Teriyaki",
                "itemid": None,
                "categoryid": None,
                "description": "Japanese style with soy glaze",
                "price": 259,
                "price_variants": [259],
                "currency": "INR",
                "instock": 2,
                "availability": 1,
                "tags": ["non-veg"],
                "addongroups": [],
                "coordinates": None,
                "confidence": 1.0
            }
        ],
        "addongroups": [
            {
                "group_name": "Extras",
                "group_id": None,
                "min_select": 0,
                "max_select": 2,
                "items": ["extra cheese"]
            }
        ],
        "audit_log": []
    }
    
    try:
        excel_path, json_path = export_menu_to_excel(
            sample_json, 
            "output",
            "sample_menu_export"
        )
        print(f"✅ Export successful!")
        print(f"Excel file: {excel_path}")
        print(f"JSON file: {json_path}")
    except Exception as e:
        print(f"❌ Export failed: {e}")