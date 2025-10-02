"""
Enhanced Excel exporter that creates multiple sheets with all data from data_reference.json
"""

import json
import logging
import os
import sys
from datetime import datetime
from pathlib import Path
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import Font, PatternFill, Alignment

# Set up logging
logging.basicConfig(
    level=logging.INFO,
    format='%(levelname)s:%(name)s:%(message)s'
)

def load_json_file(file_path: str) -> dict:
    """Load JSON file"""
    try:
        with open(file_path, 'r', encoding='utf-8') as f:
            data = json.load(f)
        logging.info(f"✅ Successfully loaded JSON from: {file_path}")
        return data
    except Exception as e:
        logging.error(f"❌ Failed to load JSON file: {e}")
        return None

def create_restaurant_sheet(data: dict) -> pd.DataFrame:
    """Create Restaurant information sheet"""
    restaurants = data.get('restaurants', [])
    rows = []
    
    for restaurant in restaurants:
        details = restaurant.get('details', {})
        rows.append({
            'Active': restaurant.get('active', ''),
            'Restaurant Name': details.get('restaurantname', ''),
            'Country': details.get('country', ''),
            'Address': details.get('address', ''),
            'Contact': details.get('contact', ''),
            'Latitude': details.get('latitude', ''),
            'Longitude': details.get('longitude', ''),
            'Cuisines': details.get('cuisines', ''),
            'Landmark': details.get('landmark', ''),
            'City': details.get('city', ''),
            'State': details.get('state', ''),
            'Menu Sharing Code': details.get('menu_sharing_code', ''),
            'Status': details.get('status', ''),
            'Table No': details.get('table_no', ''),
        })
    
    return pd.DataFrame(rows)

def create_areas_sheet(data: dict) -> pd.DataFrame:
    """Create Areas information sheet"""
    areas = data.get('areas', [])
    rows = []
    
    for area in areas:
        rows.append({
            'Restaurant Area ID': area.get('restaurantareaid', ''),
            'Area ID': area.get('areaid', ''),
            'Display Name': area.get('displayname', ''),
            'Active': area.get('active', ''),
            'Rank': area.get('rank', ''),
        })
    
    return pd.DataFrame(rows)

def create_tables_sheet(data: dict) -> pd.DataFrame:
    """Create Tables information sheet"""
    tables = data.get('tables', [])
    rows = []
    
    for table in tables:
        rows.append({
            'Table ID': table.get('id', ''),
            'Restaurant Area ID': table.get('restaurantareaid', ''),
            'Table Number': table.get('table_no', ''),
            'Active': table.get('active', ''),
            'Rank': table.get('rank', ''),
        })
    
    return pd.DataFrame(rows)

def create_categories_sheet(data: dict) -> pd.DataFrame:
    """Create Categories information sheet"""
    categories = data.get('categories', [])
    rows = []
    
    for category in categories:
        rows.append({
            'Category ID': category.get('categoryid', ''),
            'Category Name': category.get('categoryname', ''),
            'Active': category.get('active', ''),
            'Category Rank': category.get('categoryrank', ''),
            'Restaurant IDs': category.get('restarant_ids', ''),
            'Category Image URL': category.get('category_image_url', ''),
            'Parent Category ID': category.get('parent_category_id', ''),
            'Category Timings': category.get('categorytimings', ''),
        })
    
    return pd.DataFrame(rows)

def create_items_sheet(data: dict) -> pd.DataFrame:
    """Create Items information sheet (without variations and addons)"""
    items = data.get('items', [])
    rows = []
    
    for item in items:
        rows.append({
            'Item ID': item.get('itemid', ''),
            'Item Name': item.get('itemname', ''),
            'Item Rank': item.get('itemrank', ''),
            'Category ID': item.get('item_categoryid', ''),
            'Price': item.get('price', ''),
            'Active': item.get('active', ''),
            'Favorite': item.get('item_favorite', ''),
            'Allow Addon': item.get('itemallowaddon', ''),
            'Allow Variation': item.get('itemallowvariation', ''),
            'Addon Based On': item.get('itemaddonbasedon', ''),
            'In Stock': item.get('instock', ''),
            'Ignore Taxes': item.get('ignore_taxes', ''),
            'Ignore Discounts': item.get('ignore_discounts', ''),
            'Days': item.get('days', ''),
            'Attribute ID': item.get('item_attributeid', ''),
            'Description': item.get('itemdescription', ''),
            'Min Prep Time': item.get('minimumpreparationtime', ''),
            'Image URL': item.get('item_image_url', ''),
            'Tax': item.get('item_tax', ''),
            'Has Variations': len(item.get('variation', [])),
            'Has Addons': len(item.get('addon', [])),
        })
    
    return pd.DataFrame(rows)

def create_item_variations_sheet(data: dict) -> pd.DataFrame:
    """Create Item Variations sheet"""
    items = data.get('items', [])
    rows = []
    
    for item in items:
        item_id = item.get('itemid', '')
        item_name = item.get('itemname', '')
        variations = item.get('variation', [])
        
        for variation in variations:
            rows.append({
                'Item ID': item_id,
                'Item Name': item_name,
                'Variation Item ID': variation.get('variationitemid', ''),
                'Variation ID': variation.get('variationid', ''),
                'Variation Name': variation.get('variation_name', ''),
                'Variation Price': variation.get('variation_price', ''),
                'Variation Rank': variation.get('variationrank', ''),
            })
    
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        'Item ID', 'Item Name', 'Variation Item ID', 'Variation ID', 
        'Variation Name', 'Variation Price', 'Variation Rank'
    ])

def create_item_addons_sheet(data: dict) -> pd.DataFrame:
    """Create Item Addons mapping sheet"""
    items = data.get('items', [])
    rows = []
    
    for item in items:
        item_id = item.get('itemid', '')
        item_name = item.get('itemname', '')
        addons = item.get('addon', [])
        
        for addon in addons:
            rows.append({
                'Item ID': item_id,
                'Item Name': item_name,
                'Addon Group ID': addon.get('addon_group_id', ''),
                'Item Selection Type': addon.get('addon_item_selection', ''),
                'Min Selection': addon.get('addon_item_selection_min', ''),
                'Max Selection': addon.get('addon_item_selection_max', ''),
            })
    
    return pd.DataFrame(rows) if rows else pd.DataFrame(columns=[
        'Item ID', 'Item Name', 'Addon Group ID', 'Item Selection Type',
        'Min Selection', 'Max Selection'
    ])

def create_addon_groups_sheet(data: dict) -> pd.DataFrame:
    """Create Addon Groups sheet"""
    addon_groups = data.get('addongroups', [])
    rows = []
    
    for group in addon_groups:
        rows.append({
            'Addon Group ID': group.get('addongroupid', ''),
            'Addon Group Name': group.get('addongroup_name', ''),
            'Restaurant ID': group.get('addongroup_restaurantid', ''),
            'Group Rank': group.get('addongroup_rank', ''),
            'Active': group.get('active', ''),
            'Show in Online': group.get('show_in_online', ''),
            'Show in POS': group.get('show_in_pos', ''),
            'Min Quantity': group.get('min_qty', ''),
            'Max Quantity': group.get('max_qty', ''),
            'Total Items': len(group.get('addongroupitems', [])),
        })
    
    return pd.DataFrame(rows)

def create_addon_items_sheet(data: dict) -> pd.DataFrame:
    """Create Addon Items sheet"""
    addon_groups = data.get('addongroups', [])
    rows = []
    
    for group in addon_groups:
        group_id = group.get('addongroupid', '')
        group_name = group.get('addongroup_name', '')
        
        for item in group.get('addongroupitems', []):
            rows.append({
                'Addon Group ID': group_id,
                'Addon Group Name': group_name,
                'Addon Item ID': item.get('addonitemid', ''),
                'Addon Item Name': item.get('addonitem_name', ''),
                'Addon Item Price': item.get('addonitem_price', ''),
                'Active': item.get('active', ''),
                'Attributes': item.get('attributes', ''),
                'Item Rank': item.get('addonitem_rank', ''),
                'Parent Addon ID': item.get('parent_addon_id', ''),
                'Status': item.get('status', ''),
            })
    
    return pd.DataFrame(rows)

def create_complete_menu_sheet(data: dict) -> pd.DataFrame:
    """Create complete flat menu sheet with all relationships"""
    restaurant_name = data.get('restaurants', [{}])[0].get('details', {}).get('restaurantname', 'Unknown')
    
    # Create lookups
    category_map = {cat['categoryid']: cat for cat in data.get('categories', [])}
    addon_group_map = {grp['addongroupid']: grp for grp in data.get('addongroups', [])}
    area_map = {area['areaid']: area for area in data.get('areas', [])}
    
    rows = []
    items = data.get('items', [])
    
    for item in items:
        item_id = item.get('itemid', '')
        item_name = item.get('itemname', '')
        category_id = item.get('item_categoryid', '')
        category_info = category_map.get(category_id, {})
        
        base_data = {
            'Restaurant Name': restaurant_name,
            'Category ID': category_id,
            'Category Name': category_info.get('categoryname', ''),
            'Category Rank': category_info.get('categoryrank', ''),
            'Item ID': item_id,
            'Item Name': item_name,
            'Item Description': item.get('itemdescription', ''),
            'Price': item.get('price', ''),
            'Item Rank': item.get('itemrank', ''),
            'In Stock': item.get('instock', ''),
            'Active': item.get('active', ''),
            'Attribute': item.get('item_attributeid', ''),
            'Allow Addon': item.get('itemallowaddon', ''),
            'Allow Variation': item.get('itemallowvariation', ''),
        }
        
        variations = item.get('variation', [])
        addons = item.get('addon', [])
        
        if not variations and not addons:
            # Simple item
            rows.append({**base_data, 
                'Variation ID': None, 'Variation Name': None, 'Variation Price': None,
                'Addon Group ID': None, 'Addon Group Name': None, 
                'Addon Item ID': None, 'Addon Item Name': None, 'Addon Price': None
            })
        elif variations and not addons:
            # Item with variations
            for var in variations:
                rows.append({**base_data,
                    'Variation ID': var.get('variationid', ''),
                    'Variation Name': var.get('variation_name', ''),
                    'Variation Price': var.get('variation_price', ''),
                    'Addon Group ID': None, 'Addon Group Name': None,
                    'Addon Item ID': None, 'Addon Item Name': None, 'Addon Price': None
                })
        elif not variations and addons:
            # Item with addons
            for addon_ref in addons:
                group_id = addon_ref.get('addon_group_id', '')
                group_info = addon_group_map.get(group_id, {})
                
                for addon_item in group_info.get('addongroupitems', []):
                    rows.append({**base_data,
                        'Variation ID': None, 'Variation Name': None, 'Variation Price': None,
                        'Addon Group ID': group_id,
                        'Addon Group Name': group_info.get('addongroup_name', ''),
                        'Addon Item ID': addon_item.get('addonitemid', ''),
                        'Addon Item Name': addon_item.get('addonitem_name', ''),
                        'Addon Price': addon_item.get('addonitem_price', ''),
                    })
        else:
            # Item with both variations and addons
            for var in variations:
                for addon_ref in addons:
                    group_id = addon_ref.get('addon_group_id', '')
                    group_info = addon_group_map.get(group_id, {})
                    
                    for addon_item in group_info.get('addongroupitems', []):
                        rows.append({**base_data,
                            'Variation ID': var.get('variationid', ''),
                            'Variation Name': var.get('variation_name', ''),
                            'Variation Price': var.get('variation_price', ''),
                            'Addon Group ID': group_id,
                            'Addon Group Name': group_info.get('addongroup_name', ''),
                            'Addon Item ID': addon_item.get('addonitemid', ''),
                            'Addon Item Name': addon_item.get('addonitem_name', ''),
                            'Addon Price': addon_item.get('addonitem_price', ''),
                        })
    
    return pd.DataFrame(rows)

def apply_formatting(excel_path: str):
    """Apply professional formatting to Excel file"""
    try:
        wb = load_workbook(excel_path)
        
        # Define styles
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        center_alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        
        # Format each sheet
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            
            # Format header row
            if ws.max_row > 0:
                for cell in ws[1]:
                    cell.font = header_font
                    cell.fill = header_fill
                    cell.alignment = center_alignment
            
            # Auto-adjust column widths
            for column in ws.columns:
                max_length = 0
                column_letter = column[0].column_letter
                
                for cell in column:
                    try:
                        if cell.value and len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                
                adjusted_width = min(max_length + 3, 60)  # Cap at 60 characters
                ws.column_dimensions[column_letter].width = adjusted_width
        
        wb.save(excel_path)
        logging.info("✅ Applied formatting to Excel file")
        
    except Exception as e:
        logging.warning(f"⚠️  Could not apply formatting: {e}")

def create_comprehensive_excel(input_json_path: str, output_dir: str):
    """Create comprehensive Excel with multiple sheets"""
    print("🚀 Starting comprehensive Excel generation...")
    print(f"📂 Input: {input_json_path}")
    print(f"📂 Output: {output_dir}")
    print()
    
    # Load JSON data
    data = load_json_file(input_json_path)
    if not data:
        return False
    
    # Log structure
    print("📊 JSON Structure:")
    print(f"  - Restaurants: {len(data.get('restaurants', []))}")
    print(f"  - Areas: {len(data.get('areas', []))}")
    print(f"  - Tables: {len(data.get('tables', []))}")
    print(f"  - Categories: {len(data.get('categories', []))}")
    print(f"  - Items: {len(data.get('items', []))}")
    print(f"  - Addon Groups: {len(data.get('addongroups', []))}")
    print()
    
    # Create output directory
    os.makedirs(output_dir, exist_ok=True)
    
    # Generate output filename
    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    output_file = os.path.join(output_dir, f"complete_menu_data_{timestamp}.xlsx")
    
    try:
        print("📝 Creating Excel sheets...")
        
        # Create Excel writer
        with pd.ExcelWriter(output_file, engine='openpyxl') as writer:
            # Sheet 1: Restaurant Information
            print("  ✓ Creating Restaurant sheet...")
            df_restaurant = create_restaurant_sheet(data)
            df_restaurant.to_excel(writer, sheet_name='Restaurant', index=False)
            
            # Sheet 2: Areas
            print("  ✓ Creating Areas sheet...")
            df_areas = create_areas_sheet(data)
            df_areas.to_excel(writer, sheet_name='Areas', index=False)
            
            # Sheet 3: Tables
            print("  ✓ Creating Tables sheet...")
            df_tables = create_tables_sheet(data)
            df_tables.to_excel(writer, sheet_name='Tables', index=False)
            
            # Sheet 4: Categories
            print("  ✓ Creating Categories sheet...")
            df_categories = create_categories_sheet(data)
            df_categories.to_excel(writer, sheet_name='Categories', index=False)
            
            # Sheet 5: Items (master list)
            print("  ✓ Creating Items sheet...")
            df_items = create_items_sheet(data)
            df_items.to_excel(writer, sheet_name='Items', index=False)
            
            # Sheet 6: Item Variations
            print("  ✓ Creating Item Variations sheet...")
            df_variations = create_item_variations_sheet(data)
            df_variations.to_excel(writer, sheet_name='Item_Variations', index=False)
            
            # Sheet 7: Item Addons (mapping)
            print("  ✓ Creating Item Addons sheet...")
            df_item_addons = create_item_addons_sheet(data)
            df_item_addons.to_excel(writer, sheet_name='Item_Addons', index=False)
            
            # Sheet 8: Addon Groups
            print("  ✓ Creating Addon Groups sheet...")
            df_addon_groups = create_addon_groups_sheet(data)
            df_addon_groups.to_excel(writer, sheet_name='Addon_Groups', index=False)
            
            # Sheet 9: Addon Items
            print("  ✓ Creating Addon Items sheet...")
            df_addon_items = create_addon_items_sheet(data)
            df_addon_items.to_excel(writer, sheet_name='Addon_Items', index=False)
            
            # Sheet 10: Complete Menu (flat structure)
            print("  ✓ Creating Complete Menu sheet...")
            df_complete = create_complete_menu_sheet(data)
            df_complete.to_excel(writer, sheet_name='Complete_Menu', index=False)
        
        print()
        print("🎨 Applying formatting...")
        apply_formatting(output_file)
        
        print()
        print("✅ Excel file created successfully!")
        print(f"📊 File: {output_file}")
        print()
        
        # Print sheet statistics
        print("📈 Sheet Statistics:")
        print(f"  1. Restaurant: {len(df_restaurant)} rows")
        print(f"  2. Areas: {len(df_areas)} rows")
        print(f"  3. Tables: {len(df_tables)} rows")
        print(f"  4. Categories: {len(df_categories)} rows")
        print(f"  5. Items: {len(df_items)} rows")
        print(f"  6. Item_Variations: {len(df_variations)} rows")
        print(f"  7. Item_Addons: {len(df_item_addons)} rows")
        print(f"  8. Addon_Groups: {len(df_addon_groups)} rows")
        print(f"  9. Addon_Items: {len(df_addon_items)} rows")
        print(f"  10. Complete_Menu: {len(df_complete)} rows (fully expanded)")
        print()
        
        return True
        
    except Exception as e:
        logging.error(f"❌ Failed to create Excel: {e}")
        import traceback
        traceback.print_exc()
        return False

def main():
    """Main function"""
    script_dir = Path(__file__).parent
    input_path = script_dir / "sample" / "data_reference.json"
    output_dir = script_dir / "output"
    
    if not input_path.exists():
        print(f"❌ Input file not found: {input_path}")
        return
    
    success = create_comprehensive_excel(str(input_path), str(output_dir))
    
    if success:
        print("🎉 All done! Check the output directory for your comprehensive Excel file.")
    else:
        print("❌ Export failed. Check the logs above for details.")

if __name__ == "__main__":
    main()
